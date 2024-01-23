import openpyxl as op
import numpy as np

wb = op.load_workbook('model_devoir.xlsx')
ws = wb['Feuil1']
n = 10
p = 3
Y = np.array([ws.cell(2 + i, 1).value for i in range(n)])
X = np.array([[ws.cell(2 + i, 2 + j).value for j in range(p)] for i in range(n)])


colonne_un = np.ones((n, 1))

X = np.concatenate((colonne_un ,X), axis=1)
Xt = np.transpose(X)
XtX = np.dot(Xt, X)
inv = np.linalg.inv(XtX)
# a* = (inv * Xt) * y
a = np.dot(np.dot(inv, Xt), Y)

Y_ch = np.dot(X, a)

Y_bar = np.average(Y)
SCT = np.sum((Y - Y_bar) ** 2)
SCE = np.sum((Y_ch - Y_bar) ** 2)
SCR = SCT - SCE

VE = SCE/p
VR = SCR/(n - p -1)

F = VE/VR
R2 = SCE/SCT
print(R2)